from openpyxl import Workbook,load_workbook
import openpyxl
import os
import time

excel_file_name = "product.xlsx"
my_dict = {
    'name': 'xxxx',
    'ami': 'New York',
    'gender': 'xxxx',
    'coll': 'xxxx',
    'colwl': '2222',
    'coldwfl': 'efrf'
}

### Add Products Details in Excel Sheet By Dictionary
def add_data(my_dict,excel_file_name):
    wb = openpyxl.load_workbook(excel_file_name)
    ws = wb.active
    last_row = ws.max_row
    for key in my_dict.keys():
        if key in ws[1]:
            col_num = ws[1].index(key) + 1
            for row in range(2, last_row+2):
                if ws.cell(row=row, column=col_num).value is None:
                    ws.cell(row=row, column=col_num, value=my_dict[key])
                    break
        else:
            new_row = last_row + 1
            ws.cell(row=1, column=ws.max_column+1, value=key)
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == key:
                    ws.cell(row=new_row, column=col, value=my_dict[key])
                    break
    wb.save(excel_file_name)
    del_dup()


### Duplicate header delete
def del_dup(excel_file_name):
    workbook = openpyxl.load_workbook(excel_file_name)
    worksheet = workbook.active
    seen_headers = []
    for cell in worksheet[1]:
        if cell.value in seen_headers:
            worksheet.delete_cols(cell.col_idx)
        else:
            seen_headers.append(cell.value)
    workbook.save(excel_file_name)

### All duplicate value delete
wb = openpyxl.load_workbook(excel_file_name)
ws = wb.active
unique_values = set()
for row_num in range(2, ws.max_row+1):
    cell = ws.cell(row=row_num, column=1)
    if cell.value in unique_values:
        ws.delete_rows(row_num, 1)
    else:
        unique_values.add(cell.value)
wb.save(excel_file_name)

### Excel Data change
cell_row = 1
while True:
    try:
        workbook = load_workbook(filename=excel_file_name)
        spreadsheet = workbook.active
        
        URL = spreadsheet.cell(row = cell_row, column = 1).value ### Value row number

        spreadsheet.cell(row = cell_row, column = 9).value = pdf_link  ### Add value row number
        workbook.save(filename=excel_file_name)
        print(cell_row)
        cell_row += 1
    except Exception:
        time.sleep(30)
        pass
    break

#### Create Workbook File
def create_workbook(excel_file_name):
    open(excel_file_name, "a")
    if os.path.getsize(excel_file_name) == 0:
        workbook = Workbook()
        spreadsheet = workbook.active
        spreadsheet.cell(row=1, column=1).value="Title"
        spreadsheet.cell(row=1, column=2).value="Manufacture"
        spreadsheet.cell(row=1, column=3).value="ID"
        spreadsheet.cell(row=1, column=4).value="Description"
        workbook.save(filename=excel_file_name)
        
### Add Products Details in Excel Sheet
def excel_add(product_details,excel_file_name):
    workbook = load_workbook(filename=excel_file_name)
    spreadsheet = workbook.active
    ### Write
    empty_cell_row = 1
    while True:
        cell_value = spreadsheet.cell(row = empty_cell_row, column = 1).value
        if cell_value == None:
            break
        empty_cell_row += 1
    entry_column = 1
    for i in product_details:
        spreadsheet.cell(row = empty_cell_row, column = entry_column).value = i
        entry_column += 1

    workbook.save(filename=excel_file_name)

#### Duplicate value check in EXCEL Sheet
def dup_check(URL,excel_file_name):
    workbook = openpyxl.open(excel_file_name)
    spreadsheet = workbook.active
    number_rows = spreadsheet.max_row
    number_columns = spreadsheet.max_column
    for row in range(1, number_rows+ 1):
        for column in range(1, number_columns+ 1):
            cell = spreadsheet.cell(row, column)
            if cell.value == URL:
                print("Duplicate Found")
                return True
            else:
                pass
    return False